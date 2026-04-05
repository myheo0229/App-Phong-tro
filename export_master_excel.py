#!/usr/bin/env python3
"""
Export Excel updater for Room Rental App.
Cập nhật file Excel quản lý điện nước, chi phí tiền trọ từ dữ liệu app.

Cách sử dụng:
    # Cập nhật 1 phòng:
    python export_master_excel.py --room <phong> --month <thang> --year <nam> [--fullname <ten> --cccd <cccd>]
    
    # Cập nhật nhiều phòng (từ JSON):
    python export_master_excel.py --json '<json_data>'

Ví dụ:
    python export_master_excel.py --room 1A --month 10 --year 2024 --fullname "Nguyen Van A" --cccd "123456789"
    python export_master_excel.py --json '{"month":10,"year":2024,"rooms":[{"name":"1A","fullname":"Nguyen Van A","cccd":"123456789","elecUsed":100},{"name":"2A","fullname":"Tran Thi B","cccd":"987654321","elecUsed":150}]}'
"""

import argparse
import json
import sys
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Tên file Excel
EXCEL_FILE = "File Excel quản lý điện nước, chi phí tiền trọ.xlsx"
SHEET_NAME = "Quản lý điện nước, chi phí tiền"

# Các cột trong Excel (index bắt đầu từ 1)
COLUMNS = {
    'so_phong': 1,           # A: Số phòng
    'dien_cu': 2,            # B: Số công tơ điện - Số cũ
    'dien_moi': 3,           # C: Số công tơ điện - Số mới
    'dien_su_dung': 4,       # D: SỐ ĐIỆN TIÊU THỤ
    'nuoc_cu': 5,            # E: Nước - Số cũ
    'nuoc_moi': 6,           # F: Nước - Số mới
    'nuoc_su_dung': 7,       # G: SỐ NƯỚC TIÊU THỤ
    'tien_dien': 8,          # H: TIỀN ĐIỆN /SỐ
    'tien_nuoc': 9,          # I: TIỀN NƯỚC/ KHỐI
    'tien_phong': 10,        # J: TIỀN PHÒNG
    'rac': 11,               # K: RÁC
    'internet': 12,          # L: INTERNET
    'dien_hao_tai_kwh': 13,  # M: ĐIỆN HAO TẢI (Kwh)
    'tien_dien_hao_tai': 14, # N: TIỀN ĐIỆN HAO TẢI
    'no_cu': 15,             # O: NỢ CŨ
    'tong_cong': 16,         # P: TỔNG CỘNG
    'ho_ten': 17,            # Q: HỌ TÊN
    'cccd': 18,              # R: CMND
    'ki_ten': 19,            # S: KÍ TÊN
}


def find_column_by_header(ws, header_text):
    """Tìm cột theo nội dung header (có thể nằm ở row 1 hoặc row 2)."""
    for row in range(1, 4):
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value and header_text.lower() in str(cell_value).lower().strip():
                return col
    return None


def normalize_room_name(room_name):
    """Chuẩn hóa tên phòng: trim + uppercase."""
    if room_name is None:
        return ""
    return str(room_name).strip().upper()


def find_row_by_room(ws, room_name, start_row=4):
    """Tìm dòng theo tên phòng (bắt đầu từ dòng 4 - vì row 1-3 là header)."""
    normalized_room = normalize_room_name(room_name)
    for row in range(start_row, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=COLUMNS['so_phong']).value
        if normalize_room_name(cell_value) == normalized_room:
            return row
    return None


def ensure_electricity_loss_columns(ws):
    """Đảm bảo cột 'Điện hao tải (kWh)' và 'Tiền điện hao tải' tồn tại.
    Nếu chưa có thì thêm vào."""
    # Kiểm tra cột ĐIỆN HAO TẢI (Kwh) đã tồn tại chưa
    dien_hao_tai_col = find_column_by_header(ws, "điện hao tải")
    tien_dien_hao_tai_col = find_column_by_header(ws, "tiền điện hao tải")
    
    # Nếu chưa có cột điện hao tải, cần thêm
    if dien_hao_tai_col is None:
        # Tìm vị trí chèn (trước cột TỔNG CỘNG)
        tong_cong_col = find_column_by_header(ws, "tổng cộng")
        if tong_cong_col is None:
            tong_cong_col = COLUMNS['tong_cong']
        
        # Chèn 2 cột mới trước cột TỔNG CỘNG
        insert_col = tong_cong_col
        
        # Chèn cột TIỀN ĐIỆN HAO TẢI
        ws.insert_cols(insert_col)
        ws.cell(row=2, column=insert_col).value = "TIỀN ĐIỆN HAO TẢI"
        ws.cell(row=3, column=insert_col).value = "Thành tiền"
        
        # Chèn cột ĐIỆN HAO TẢI (Kwh)
        ws.insert_cols(insert_col)
        ws.cell(row=2, column=insert_col).value = "ĐIỆN HAO TẢI (Kwh)"
        ws.cell(row=3, column=insert_col).value = "Số kWh hao tải"
        
        # Cập nhật lại các column index
        update_column_indices_after_insert(ws, insert_col, 2)
        
        # Thêm công thức cho các dòng dữ liệu (bắt đầu từ row 4)
        for row in range(4, ws.max_row + 1):
            room_cell = ws.cell(row=row, column=COLUMNS['so_phong']).value
            if room_cell and normalize_room_name(room_cell):
                # ĐIỆN HAO TẢI (Kwh) = SỐ ĐIỆN TIÊU THỤ * 7%
                ws.cell(row=row, column=insert_col).value = f"=D{row}*7%"
                # TIỀN ĐIỆN HAO TẢI = ĐIỆN HAO TẢI * 2900
                ws.cell(row=row, column=insert_col + 1).value = f"=ROUND({get_column_letter(insert_col)}{row}*2900, -3)"
        
        return True
    
    return False


def update_column_indices_after_insert(ws, insert_col, num_cols_inserted):
    """Cập nhật lại các index cột sau khi chèn."""
    for key, col in COLUMNS.items():
        if col >= insert_col:
            COLUMNS[key] = col + num_cols_inserted


def update_room_data(ws, row, fullname=None, cccd=None, elec_used=None,
                     electricity_loss_percent=7, electricity_price=2900):
    """Cập nhật thông tin cho 1 phòng.
    
    Args:
        ws: Worksheet
        row: Dòng cần cập nhật
        fullname: Họ tên (nếu có thì ghi đè, không có thì giữ nguyên)
        cccd: CCCD (nếu có thì ghi đè, không có thì giữ nguyên)
        elec_used: Số điện đã sử dụng (kWh) - để tính hao tải
        electricity_loss_percent: % hao tải
        electricity_price: Giá điện cho hao tải
    """
    # Cập nhật thông tin người thuê nếu có
    if fullname is not None and str(fullname).strip():
        ws.cell(row=row, column=COLUMNS['ho_ten']).value = str(fullname).strip()
    
    if cccd is not None and str(cccd).strip():
        ws.cell(row=row, column=COLUMNS['cccd']).value = str(cccd).strip()
    
    # Cập nhật công thức ĐIỆN HAO TẢI (Kwh) và TIỀN ĐIỆN HAO TẢI
    dien_hao_tai_col = find_column_by_header(ws, "điện hao tải")
    tien_dien_hao_tai_col = find_column_by_header(ws, "tiền điện hao tải")
    
    if dien_hao_tai_col and tien_dien_hao_tai_col:
        # Cập nhật công thức cho cột ĐIỆN HAO TẢI (Kwh)
        ws.cell(row=row, column=dien_hao_tai_col).value = f"=D{row}*{electricity_loss_percent}%"
        # Cập nhật công thức cho cột TIỀN ĐIỆN HAO TẢI
        ws.cell(row=row, column=tien_dien_hao_tai_col).value = f"=ROUND({get_column_letter(dien_hao_tai_col)}{row}*{electricity_price}, -3)"
    
    # Cập nhật công thức TỔNG CỘNG
    tong_cong_col = find_column_by_header(ws, "tổng cộng")
    if tong_cong_col:
        # Tổng = TIỀN ĐIỆN + TIỀN NƯỚC + TIỀN PHÒNG + RÁC + INTERNET + TIỀN ĐIỆN HAO TẢI + NỢ CŨ
        ws.cell(row=row, column=tong_cong_col).value = f"=ROUND(SUM(H{row}:{get_column_letter(tong_cong_col-1)}{row}), -3)"
    
    # KHÔNG cập nhật cột KÍ TÊN (column S)


def update_excel_batch(rooms_data, month, year, 
                       electricity_loss_percent=7, electricity_price=2900):
    """Cập nhật Excel cho nhiều phòng cùng lúc.
    
    Args:
        rooms_data: List of dict với keys: name, fullname (optional), cccd (optional), elecUsed (optional)
        month: Tháng
        year: Năm
        electricity_loss_percent: % hao tải điện
        electricity_price: Giá điện cho hao tải
    
    Returns:
        dict: Kết quả cập nhật
    """
    # Kiểm tra file tồn tại
    if not os.path.exists(EXCEL_FILE):
        return {
            'success': False,
            'error': f"Không tìm thấy file Excel: {EXCEL_FILE}"
        }
    
    try:
        # Load workbook
        wb = load_workbook(EXCEL_FILE)
        
        # Kiểm tra sheet tồn tại
        if SHEET_NAME not in wb.sheetnames:
            return {
                'success': False,
                'error': f"Không tìm thấy sheet: {SHEET_NAME}"
            }
        
        ws = wb[SHEET_NAME]
        
        # Đảm bảo các cột điện hao tải tồn tại
        ensure_electricity_loss_columns(ws)
        
        updated_rooms = []
        not_found_rooms = []
        
        for room_data in rooms_data:
            room_name = room_data.get('name')
            if not room_name:
                continue
            
            # Tìm dòng theo phòng
            row = find_row_by_room(ws, room_name)
            
            if row is None:
                not_found_rooms.append(room_name)
                continue
            
            # Cập nhật thông tin cho phòng
            update_room_data(
                ws, row,
                fullname=room_data.get('fullname'),
                cccd=room_data.get('cccd'),
                elec_used=room_data.get('elecUsed'),
                electricity_loss_percent=electricity_loss_percent,
                electricity_price=electricity_price
            )
            
            updated_rooms.append(room_name)
        
        # Lưu file
        wb.save(EXCEL_FILE)
        wb.close()
        
        return {
            'success': True,
            'message': f"Đã cập nhật {len(updated_rooms)} phòng",
            'updated_rooms': updated_rooms,
            'not_found_rooms': not_found_rooms,
            'month': month,
            'year': year
        }
        
    except Exception as e:
        return {
            'success': False,
            'error': f"Lỗi khi cập nhật Excel: {str(e)}"
        }


def update_excel(room, month, year, fullname=None, cccd=None, 
                 electricity_loss_percent=7, electricity_price=2900):
    """Hàm cập nhật 1 phòng (wrapper cho update_excel_batch)."""
    rooms_data = [{
        'name': room,
        'fullname': fullname,
        'cccd': cccd
    }]
    
    result = update_excel_batch(
        rooms_data, month, year,
        electricity_loss_percent, electricity_price
    )
    
    if result['success']:
        result['fullname_updated'] = fullname is not None and str(fullname).strip() != ''
        result['cccd_updated'] = cccd is not None and str(cccd).strip() != ''
    
    return result


def main():
    parser = argparse.ArgumentParser(description='Cập nhật file Excel quản lý phòng trọ')
    parser.add_argument('--room', help='Tên phòng (ví dụ: 1A)')
    parser.add_argument('--month', type=int, help='Tháng (1-12)')
    parser.add_argument('--year', type=int, help='Năm (ví dụ: 2024)')
    parser.add_argument('--fullname', help='Họ tên người thuê')
    parser.add_argument('--cccd', help='Số CCCD')
    parser.add_argument('--electricity-loss', type=float, default=7, 
                        help='Tỷ lệ hao tải điện (%)')
    parser.add_argument('--electricity-price', type=int, default=2900,
                        help='Giá điện cho hao tải (đ/kWh)')
    parser.add_argument('--json', help='JSON data với danh sách phòng cần cập nhật')
    
    args = parser.parse_args()
    
    # Chế độ batch từ JSON
    if args.json:
        try:
            data = json.loads(args.json)
            rooms_data = data.get('rooms', [])
            month = data.get('month', args.month)
            year = data.get('year', args.year)
            
            if not month or not year:
                print("❌ Lỗi: Cần cung cấp month và year")
                sys.exit(1)
            
            result = update_excel_batch(
                rooms_data, month, year,
                args.electricity_loss, args.electricity_price
            )
            
            if result['success']:
                print(f"✅ {result['message']}")
                for room in result.get('updated_rooms', []):
                    print(f"   - {room}")
                if result.get('not_found_rooms'):
                    print(f"   Không tìm thấy: {', '.join(result['not_found_rooms'])}")
            else:
                print(f"❌ Lỗi: {result.get('error', 'Không rõ lỗi')}")
                sys.exit(1)
                
        except json.JSONDecodeError as e:
            print(f"❌ Lỗi JSON: {e}")
            sys.exit(1)
        return
    
    # Chế độ 1 phòng
    if not args.room or not args.month or not args.year:
        print("❌ Lỗi: Cần cung cấp --room, --month, --year hoặc --json")
        sys.exit(1)
    
    result = update_excel(
        room=args.room,
        month=args.month,
        year=args.year,
        fullname=args.fullname,
        cccd=args.cccd,
        electricity_loss_percent=args.electricity_loss,
        electricity_price=args.electricity_price
    )
    
    if result['success']:
        print(f"✅ {result['message']}")
        if result.get('fullname_updated'):
            print(f"   - Đã cập nhật họ tên: {args.fullname}")
        if result.get('cccd_updated'):
            print(f"   - Đã cập nhật CCCD: {args.cccd}")
    else:
        print(f"❌ Lỗi: {result.get('error', 'Không rõ lỗi')}")
        sys.exit(1)


if __name__ == "__main__":
    main()